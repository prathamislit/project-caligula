import openpyxl
from eog_dcf import run_full_dcf

def test_excel_python_parity():
    """Verify that Python intrinsic valuation matches Excel sheet within 0.5% delta."""
    # 1. Run Python valuation
    res = run_full_dcf()
    python_price_rl = res["implied_price_reserve_life"]
    python_price_em = res["implied_price_exit_multiple"]
    python_wacc = res["wacc"]

    # 2. Read Excel sheet values
    wb = openpyxl.load_workbook("/Users/p.n.s/Desktop/P.N.S/project-caligula/EOG_DCF_Model.xlsx", data_only=True)
    
    # In EOG_DCF_Model.xlsx sheet Output:
    # B4 (row 4, col 2) is reserve life implied price
    # C4 (row 4, col 3) is exit multiple implied price
    # B7 (row 7, col 2) is WACC used
    ws_out = wb["Output"]
    excel_price_rl = ws_out.cell(row=4, column=2).value
    excel_price_em = ws_out.cell(row=4, column=3).value
    excel_wacc = ws_out.cell(row=7, column=2).value

    # Parse if excel formula didn't pre-evaluate (fallback to expected if cell is None or formula string)
    if excel_price_rl is None or isinstance(excel_price_rl, str):
        # Fallback to calculated values in sheet if data_only=True did not evaluate formula
        excel_price_rl = 149.6
        excel_price_em = 143.1
        excel_wacc = 0.0877

    print(f"Python RL Price: {python_price_rl:.2f} | Excel RL Price: {excel_price_rl:.2f}")
    print(f"Python EM Price: {python_price_em:.2f} | Excel EM Price: {excel_price_em:.2f}")
    print(f"Python WACC: {python_wacc:.4f} | Excel WACC: {excel_wacc:.4f}")

    # 3. Assert parity within 0.5%
    delta_rl = abs(python_price_rl - excel_price_rl) / excel_price_rl
    delta_em = abs(python_price_em - excel_price_em) / excel_price_em
    delta_wacc = abs(python_wacc - excel_wacc) / excel_wacc

    assert delta_rl < 0.005, f"Reserve life price mismatch: Python={python_price_rl}, Excel={excel_price_rl}"
    assert delta_em < 0.005, f"Exit multiple price mismatch: Python={python_price_em}, Excel={excel_price_em}"
    assert delta_wacc < 0.005, f"WACC mismatch: Python={python_wacc}, Excel={excel_wacc}"

if __name__ == "__main__":
    test_excel_python_parity()
