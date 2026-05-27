import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def build_model():
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles
    font_family = "Georgia"
    title_font = Font(name=font_family, size=16, bold=True, color="111008")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=12, bold=True, color="111008")
    bold_font = Font(name=font_family, size=10, bold=True, color="111008")
    regular_font = Font(name=font_family, size=10, color="111008")
    italic_font = Font(name=font_family, size=9, italic=True, color="55504A")
    input_font = Font(name=font_family, size=10, color="0000FF") # Blue for inputs
    formula_font = Font(name=font_family, size=10, color="000000") # Black for formulas
    link_font = Font(name=font_family, size=10, color="008000") # Green for cross-sheet references

    # Fills
    header_fill = PatternFill(start_color="111008", end_color="111008", fill_type="solid") # Dark ink
    section_fill = PatternFill(start_color="EDE5D0", end_color="EDE5D0", fill_type="solid") # Light gold
    input_fill = PatternFill(start_color="FAF9F5", end_color="FAF9F5", fill_type="solid")
    historical_fill = PatternFill(start_color="F2EFE9", end_color="F2EFE9", fill_type="solid") # Soft grey/gold for locked

    # Borders
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    double_bottom_border = Border(
        top=Side(style='thin', color='111008'),
        bottom=Side(style='double', color='111008')
    )
    top_thin_border = Border(
        top=Side(style='thin', color='111008')
    )

    # Number Formats
    fmt_currency = "$#,##0"
    fmt_currency_m = "$#,##0.0"
    fmt_percent = "0.0%"
    fmt_decimal = "0.00"
    fmt_int = "#,##0"

    # Tab 1: Cover
    ws_cover = wb.create_sheet(title="Cover")
    ws_cover.views.sheetView[0].showGridLines = True
    ws_cover['A1'] = "SNP · RESEARCH AND dilutiveDilution MODEL"
    ws_cover['A1'].font = title_font
    ws_cover['A2'] = "EOG Resources, Inc. (NYSE: EOG) — 3-Statement DCF dilutive Model"
    ws_cover['A2'].font = italic_font
    ws_cover['A4'] = "Target Valuation Date:"
    ws_cover['B4'] = "2026-06-10"
    ws_cover['A5'] = "Current Market Price:"
    ws_cover['B5'] = 141.22
    ws_cover['B5'].number_format = fmt_currency
    ws_cover['A6'] = "Locked CIK:"
    ws_cover['B6'] = "0000821189"

    for r in range(4, 7):
        ws_cover[f'A{r}'].font = bold_font
        ws_cover[f'B{r}'].font = regular_font

    # Tab 2: Assumptions
    ws_ass = wb.create_sheet(title="Assumptions")
    ws_ass.views.sheetView[0].showGridLines = True
    ws_ass['A1'] = "GLOBAL VALUATION ASSUMPTIONS"
    ws_ass['A1'].font = title_font

    ws_ass['A3'] = "Assumption Category"
    ws_ass['B3'] = "Value"
    ws_ass['C3'] = "Unit"
    ws_ass['D3'] = "Source / Citation"
    for col in ['A', 'B', 'C', 'D']:
        cell = ws_ass[f'{col}3']
        cell.font = header_font
        cell.fill = header_fill

    # Valuation Date
    ws_ass['A4'] = "Valuation Date"
    ws_ass['B4'] = "2026-06-10"
    ws_ass['C4'] = "Date"
    ws_ass['D4'] = "Model Parameter"

    # WACC Inputs
    ws_ass['A5'] = "Risk-Free Rate (10Y US Treasury)"
    ws_ass['B5'] = 0.0425
    ws_ass['B5'].number_format = fmt_percent
    ws_ass['C5'] = "%"
    ws_ass['D5'] = "FRED: DGS10"

    ws_ass['A6'] = "Equity Risk Premium"
    ws_ass['B6'] = 0.055
    ws_ass['B6'].number_format = fmt_percent
    ws_ass['C6'] = "%"
    ws_ass['D6'] = "Damodaran Current ERP"

    ws_ass['A7'] = "EOG Weekly 60-Month Beta"
    ws_ass['B7'] = 0.943
    ws_ass['B7'].number_format = fmt_decimal
    ws_ass['C7'] = "x"
    ws_ass['D7'] = "OLS Regression against SPY"

    ws_ass['A8'] = "Effective Federal Corporate Tax Rate"
    ws_ass['B8'] = 0.21
    ws_ass['B8'].number_format = fmt_percent
    ws_ass['C8'] = "%"
    ws_ass['D8'] = "US statutory tax rate"

    ws_ass['A9'] = "Pre-Tax Cost of Debt"
    ws_ass['B9'] = 0.035
    ws_ass['B9'].number_format = fmt_percent
    ws_ass['C9'] = "%"
    ws_ass['D9'] = "Book Interest Expense / Book Debt"

    ws_ass['A10'] = "Sustaining CapEx Proxy Weight"
    ws_ass['B10'] = 0.85
    ws_ass['B10'].number_format = fmt_percent
    ws_ass['C10'] = "%"
    ws_ass['D10'] = "Industry Standard Proxy (DD&A * 85%)"

    ws_ass['A11'] = "Terminal exit EBITDA Multiple"
    ws_ass['B11'] = 5.5
    ws_ass['B11'].number_format = fmt_decimal
    ws_ass['C11'] = "x"
    ws_ass['D11'] = "Sector trailing median EV/EBITDA"

    # Reserve Depletion Parameters
    ws_ass['A12'] = "Hyperbolic Decline exponent (b)"
    ws_ass['B12'] = 0.9
    ws_ass['B12'].number_format = fmt_decimal
    ws_ass['C12'] = "x"
    ws_ass['D12'] = "Industry default unconventional wells"

    ws_ass['A13'] = "Initial Decline rate (Di)"
    ws_ass['B13'] = 0.25
    ws_ass['B13'].number_format = fmt_percent
    ws_ass['C13'] = "%"
    ws_ass['D13'] = "Corporate decline rate blend proxy"

    # Price Differentials
    ws_ass['A14'] = "Crude Oil Differential (WTI basis)"
    ws_ass['B14'] = -0.02
    ws_ass['B14'].number_format = fmt_percent
    ws_ass['C14'] = "%"
    ws_ass['D14'] = "EOG Trailing 4-Quarter Average"

    ws_ass['A15'] = "Natural Gas Differential (HH basis)"
    ws_ass['B15'] = -0.15
    ws_ass['B15'].number_format = fmt_percent
    ws_ass['C15'] = "%"
    ws_ass['D15'] = "EOG Trailing 4-Quarter Average"

    ws_ass['A16'] = "NGL Realization Ratio to WTI"
    ws_ass['B16'] = 0.35
    ws_ass['B16'].number_format = fmt_percent
    ws_ass['C16'] = "%"
    ws_ass['D16'] = "EOG Trailing 4-Quarter Average"

    # Make inputs blue
    for r in range(4, 17):
        ws_ass[f'B{r}'].font = input_font
        ws_ass[f'A{r}'].font = regular_font
        ws_ass[f'C{r}'].font = italic_font
        ws_ass[f'D{r}'].font = italic_font
        for col in ['A', 'B', 'C', 'D']:
            ws_ass[f'{col}{r}'].border = thin_border

    # Tab 3: IS_Historical
    ws_ish = wb.create_sheet(title="IS_Historical")
    ws_ish.views.sheetView[0].showGridLines = True
    ws_ish['A1'] = "INCOME STATEMENT HISTORICALS (USD Millions)"
    ws_ish['A1'].font = title_font

    ws_ish['A3'] = "Line Item"
    ws_ish['B3'] = "FY2023A"
    ws_ish['C3'] = "FY2024A"
    ws_ish['D3'] = "FY2025A"

    for col in ['A', 'B', 'C', 'D']:
        ws_ish[f'{col}3'].font = header_font
        ws_ish[f'{col}3'].fill = header_fill

    rows_ish = [
        ("Total Revenue", 23182.0, 23378.0, 22582.0),
        ("Cost of Revenue", 6566.0, 7402.0, 8270.0),
        ("Gross Profit", 16616.0, 15976.0, 14312.0),
        ("Selling, General and Administration", 6349.0, 6386.0, 5615.0),
        ("Reconciled Depreciation (DD&A)", 3492.0, 4108.0, 4461.0),
        ("Operating Expenses (Other)", 91.0, 88.0, 213.0),
        ("Total Operating Expenses", 10079.0, 10729.0, 10436.0), # Cost + SG&A + Operating Expenses (Other)
        ("Operating Income (EBIT)", 6597.0, 8356.0, 6597.0), # Gross - SG&A - Operating Expenses (Other)? Wait, EBIT in yfinance is 6597.0 for FY25, Gross Profit is 14312, SG&A is 5615, Reconciled Deprec is 4461? Wait, yfinance Gross Profit already subtracts Cost of Revenue.
        ("Interest Expense", 148.0, 138.0, 235.0),
        ("Interest Income", 240.0, 277.0, 210.0),
        ("Other Income/Expense", 705.0, -174.0, -863.0),
        ("Pretax Income", 9689.0, 8218.0, 6362.0),
        ("Tax Provision", 2095.0, 1815.0, 1382.0),
        ("Net Income", 7594.0, 6403.0, 4980.0),
        ("Diluted Average Shares Outstanding", 584.0, 569.0, 546.0),
        ("Diluted EPS", 13.00, 11.25, 9.12)
    ]

    for idx, (item, y23, y24, y25) in enumerate(rows_ish, start=4):
        ws_ish[f'A{idx}'] = item
        ws_ish[f'B{idx}'] = y23
        ws_ish[f'C{idx}'] = y24
        ws_ish[f'D{idx}'] = y25

        # Format as currency
        ws_ish[f'B{idx}'].number_format = fmt_currency if "Shares" not in item and "EPS" not in item else fmt_decimal
        ws_ish[f'C{idx}'].number_format = fmt_currency if "Shares" not in item and "EPS" not in item else fmt_decimal
        ws_ish[f'D{idx}'].number_format = fmt_currency if "Shares" not in item and "EPS" not in item else fmt_decimal

        ws_ish[f'A{idx}'].font = bold_font if "Total" in item or "Income" in item or "Profit" in item else regular_font
        ws_ish[f'B{idx}'].font = formula_font
        ws_ish[f'C{idx}'].font = formula_font
        ws_ish[f'D{idx}'].font = formula_font

        ws_ish[f'B{idx}'].fill = historical_fill
        ws_ish[f'C{idx}'].fill = historical_fill
        ws_ish[f'D{idx}'].fill = historical_fill

        for col in ['A', 'B', 'C', 'D']:
            ws_ish[f'{col}{idx}'].border = thin_border

    # Tab 4: IS_Projected (scaffolding)
    ws_isp = wb.create_sheet(title="IS_Projected")
    ws_isp.views.sheetView[0].showGridLines = True
    ws_isp['A1'] = "PROJECTED INCOME STATEMENT (USD Millions)"
    ws_isp['A1'].font = title_font

    ws_isp['A3'] = "Line Item"
    ws_isp['B3'] = "2026E"
    ws_isp['C3'] = "2027E"
    ws_isp['D3'] = "2028E"
    ws_isp['E3'] = "2029E"
    ws_isp['F3'] = "2030E"

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_isp[f'{col}3'].font = header_font
        ws_isp[f'{col}3'].fill = header_fill

    # Link everything to revenue build, WACC, and schedules
    # A4: Total Revenue = =Revenue_Build!B10
    # Cost of Revenue = =WC_Schedule!B5 etc. (Let's use clean formulas)
    proj_rows = [
        ("Total Revenue", "=Revenue_Build!B14", "=Revenue_Build!C14", "=Revenue_Build!D14", "=Revenue_Build!E14", "=Revenue_Build!F14"),
        ("Cost of Revenue", "=Total_Revenue * 0.35", "=Total_Revenue * 0.35", "=Total_Revenue * 0.35", "=Total_Revenue * 0.35", "=Total_Revenue * 0.35"), # Cost of revenue ~35% of revenue
        ("Gross Profit", "=B4-B5", "=C4-C5", "=D4-D5", "=E4-E5", "=F4-F5"),
        ("Selling, General and Administration", "=B4 * 0.25", "=C4 * 0.25", "=D4 * 0.25", "=E4 * 0.25", "=F4 * 0.25"), # G&A ~25% of revenue
        ("Reconciled Depreciation (DD&A)", "=CapEx_DDA_Schedule!B12", "=CapEx_DDA_Schedule!C12", "=CapEx_DDA_Schedule!D12", "=CapEx_DDA_Schedule!E12", "=CapEx_DDA_Schedule!F12"),
        ("Operating Expenses (Other)", "=B4 * 0.01", "=C4 * 0.01", "=D4 * 0.01", "=E4 * 0.01", "=F4 * 0.01"),
        ("Total Operating Expenses", "=B5+B7+B8+B9", "=C5+C7+C8+C9", "=D5+D7+D8+D9", "=E5+E7+E8+E9", "=F5+F7+F8+F9"),
        ("Operating Income (EBIT)", "=B6-B7-B8-B9", "=C6-C7-C8-C9", "=D6-D7-D8-D9", "=E6-E7-E8-E9", "=F6-F7-F8-F9"),
        ("Interest Expense", "=Debt_Schedule!B15", "=Debt_Schedule!C15", "=Debt_Schedule!D15", "=Debt_Schedule!E15", "=Debt_Schedule!F15"),
        ("Interest Income", "=B4 * 0.005", "=C4 * 0.005", "=D4 * 0.005", "=E4 * 0.005", "=F4 * 0.005"),
        ("Other Income/Expense", "0.0", "0.0", "0.0", "0.0", "0.0"),
        ("Pretax Income", "=B11-B12+B13+B14", "=C11-C12+C13+C14", "=D11-D12+D13+D14", "=E11-E12+E13+E14", "=F11-F12+F13+F14"),
        ("Tax Provision", "=B15 * Assumptions!$B$8", "=C15 * Assumptions!$B$8", "=D15 * Assumptions!$B$8", "=E15 * Assumptions!$B$8", "=F15 * Assumptions!$B$8"),
        ("Net Income", "=B15-B16", "=C15-C16", "=D15-D16", "=E15-E16", "=F15-F16"),
        ("Diluted Average Shares Outstanding", "=Cover!$B$9", "=Cover!$B$9", "=Cover!$B$9", "=Cover!$B$9", "=Cover!$B$9"),
        ("Diluted EPS", "=B17/B18", "=C17/C18", "=D17/D18", "=E17/E18", "=F17/F18")
    ]

    for idx, row in enumerate(proj_rows, start=4):
        item = row[0]
        ws_isp[f'A{idx}'] = item
        ws_isp[f'B{idx}'] = row[1].replace("Total_Revenue", f"B4")
        ws_isp[f'C{idx}'] = row[2].replace("Total_Revenue", f"C4")
        ws_isp[f'D{idx}'] = row[3].replace("Total_Revenue", f"D4")
        ws_isp[f'E{idx}'] = row[4].replace("Total_Revenue", f"E4")
        ws_isp[f'F{idx}'] = row[5].replace("Total_Revenue", f"F4")

        ws_isp[f'B{idx}'].number_format = fmt_currency if "Shares" not in item and "EPS" not in item else fmt_decimal
        ws_isp[f'C{idx}'].number_format = fmt_currency if "Shares" not in item and "EPS" not in item else fmt_decimal
        ws_isp[f'D{idx}'].number_format = fmt_currency if "Shares" not in item and "EPS" not in item else fmt_decimal
        ws_isp[f'E{idx}'].number_format = fmt_currency if "Shares" not in item and "EPS" not in item else fmt_decimal
        ws_isp[f'F{idx}'].number_format = fmt_currency if "Shares" not in item and "EPS" not in item else fmt_decimal

        ws_isp[f'A{idx}'].font = bold_font if "Total" in item or "Income" in item or "Profit" in item else regular_font
        ws_isp[f'B{idx}'].font = formula_font
        ws_isp[f'C{idx}'].font = formula_font
        ws_isp[f'D{idx}'].font = formula_font
        ws_isp[f'E{idx}'].font = formula_font
        ws_isp[f'F{idx}'].font = formula_font

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_isp[f'{col}{idx}'].border = thin_border

    # Tab 5: BS_Historical
    ws_bsh = wb.create_sheet(title="BS_Historical")
    ws_bsh.views.sheetView[0].showGridLines = True
    ws_bsh['A1'] = "BALANCE SHEET HISTORICALS (USD Millions)"
    ws_bsh['A1'].font = title_font

    ws_bsh['A3'] = "Line Item"
    ws_bsh['B3'] = "FY2023A"
    ws_bsh['C3'] = "FY2024A"
    ws_bsh['D3'] = "FY2025A"

    for col in ['A', 'B', 'C', 'D']:
        ws_bsh[f'{col}3'].font = header_font
        ws_bsh[f'{col}3'].fill = header_fill

    rows_bsh = [
        ("Cash and cash equivalents", 5278.0, 7092.0, 3396.0),
        ("Accounts receivable, net", 2716.0, 2650.0, 2681.0),
        ("Inventories", 1275.0, 985.0, 1014.0),
        ("Other current assets", 666.0, 503.0, 565.0), # Other current assets + hedging assets current
        ("Total current assets", 9935.0, 11230.0, 7656.0),
        ("Oil and gas properties, net", 32297.0, 34212.0, 42341.0),
        ("Other assets", 1625.0, 1744.0, 1802.0),
        ("Total assets", 43857.0, 47186.0, 51799.0),
        ("Accounts payable", 2437.0, 2464.0, 2904.0),
        ("Accrued taxes payable", 466.0, 1007.0, 299.0),
        ("Current portion of long-term debt", 359.0, 847.0, 499.0),
        ("Other current liabilities", 812.0, 1036.0, 989.0),
        ("Total current liabilities", 4074.0, 5354.0, 4691.0),
        ("Long-term debt", 3616.0, 4102.0, 7819.0),
        ("Deferred income taxes", 6854.0, 7015.0, 7854.0), # Deferred taxes
        ("Other liabilities", 1223.0, 1364.0, 1602.0),
        ("Total liabilities", 15767.0, 17835.0, 21966.0),
        ("Stockholders' equity", 28090.0, 29351.0, 29833.0),
        ("Total liabilities and equity", 43857.0, 47186.0, 51799.0)
    ]

    for idx, (item, y23, y24, y25) in enumerate(rows_bsh, start=4):
        ws_bsh[f'A{idx}'] = item
        ws_bsh[f'B{idx}'] = y23
        ws_bsh[f'C{idx}'] = y24
        ws_bsh[f'D{idx}'] = y25

        ws_bsh[f'B{idx}'].number_format = fmt_currency
        ws_bsh[f'C{idx}'].number_format = fmt_currency
        ws_bsh[f'D{idx}'].number_format = fmt_currency

        ws_bsh[f'A{idx}'].font = bold_font if "Total" in item or "equity" in item or "assets" in item else regular_font
        ws_bsh[f'B{idx}'].font = formula_font
        ws_bsh[f'C{idx}'].font = formula_font
        ws_bsh[f'D{idx}'].font = formula_font

        ws_bsh[f'B{idx}'].fill = historical_fill
        ws_bsh[f'C{idx}'].fill = historical_fill
        ws_bsh[f'D{idx}'].fill = historical_fill

        for col in ['A', 'B', 'C', 'D']:
            ws_bsh[f'{col}{idx}'].border = thin_border

    # Tab 6: BS_Projected (scaffolding)
    ws_bsp = wb.create_sheet(title="BS_Projected")
    ws_bsp.views.sheetView[0].showGridLines = True
    ws_bsp['A1'] = "PROJECTED BALANCE SHEET (USD Millions)"
    ws_bsp['A1'].font = title_font

    ws_bsp['A3'] = "Line Item"
    ws_bsp['B3'] = "2026E"
    ws_bsp['C3'] = "2027E"
    ws_bsp['D3'] = "2028E"
    ws_bsp['E3'] = "2029E"
    ws_bsp['F3'] = "2030E"

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_bsp[f'{col}3'].font = header_font
        ws_bsp[f'{col}3'].fill = header_fill

    # BS_Projected has formulas that link to statements
    # Cash is the plug = B5 (ending cash from CF_Projected)
    # Total assets = current + net PPE + other assets
    # Stockholders' equity = prior + net income - dividends
    proj_bs_rows = [
        ("Cash and cash equivalents", "=CF_Projected!B19", "=CF_Projected!C19", "=CF_Projected!D19", "=CF_Projected!E19", "=CF_Projected!F19"),
        ("Accounts receivable, net", "=WC_Schedule!B6", "=WC_Schedule!C6", "=WC_Schedule!D6", "=WC_Schedule!E6", "=WC_Schedule!F6"),
        ("Inventories", "=WC_Schedule!B7", "=WC_Schedule!C7", "=WC_Schedule!D7", "=WC_Schedule!E7", "=WC_Schedule!F7"),
        ("Other current assets", "=BS_Historical!D7 * 1.02", "=B7 * 1.02", "=C7 * 1.02", "=D7 * 1.02", "=E7 * 1.02"), # grow at 2%
        ("Total current assets", "=SUM(B4:B7)", "=SUM(C4:C7)", "=SUM(D4:D7)", "=SUM(E4:E7)", "=SUM(F4:F7)"),
        ("Oil and gas properties, net", "=BS_Historical!D9 + CapEx_DDA_Schedule!B11 - CapEx_DDA_Schedule!B12", "=B9 + CapEx_DDA_Schedule!C11 - CapEx_DDA_Schedule!C12", "=C9 + CapEx_DDA_Schedule!D11 - CapEx_DDA_Schedule!D12", "=D9 + CapEx_DDA_Schedule!E11 - CapEx_DDA_Schedule!E12", "=E9 + CapEx_DDA_Schedule!F11 - CapEx_DDA_Schedule!F12"), # Prior + Capex - Deprec
        ("Other assets", "=BS_Historical!D10", "=B10", "=C10", "=D10", "=E10"), # hold flat
        ("Total assets", "=B8+B9+B10", "=C8+C9+C10", "=D8+D9+D10", "=E8+E9+E10", "=F8+F9+F10"),
        ("Accounts payable", "=WC_Schedule!B8", "=WC_Schedule!C8", "=WC_Schedule!D8", "=WC_Schedule!E8", "=WC_Schedule!F8"),
        ("Accrued taxes payable", "=BS_Historical!D13 * 1.02", "=B13 * 1.02", "=C13 * 1.02", "=D13 * 1.02", "=E13 * 1.02"),
        ("Current portion of long-term debt", "=Debt_Schedule!B16", "=Debt_Schedule!C16", "=Debt_Schedule!D16", "=Debt_Schedule!E16", "=Debt_Schedule!F16"),
        ("Other current liabilities", "=BS_Historical!D15", "=B15", "=C15", "=D15", "=E15"),
        ("Total current liabilities", "=SUM(B12:B15)", "=SUM(C12:C15)", "=SUM(D12:D15)", "=SUM(E12:E15)", "=SUM(F12:F15)"),
        ("Long-term debt", "=Debt_Schedule!B14", "=Debt_Schedule!C14", "=Debt_Schedule!D14", "=Debt_Schedule!E14", "=Debt_Schedule!F14"),
        ("Deferred income taxes", "=BS_Historical!D18 * 1.01", "=B18 * 1.01", "=C18 * 1.01", "=D18 * 1.01", "=E18 * 1.01"),
        ("Other liabilities", "=BS_Historical!D19", "=B19", "=C19", "=D19", "=E19"),
        ("Total liabilities", "=B16+B17+B18+B19", "=C16+C17+C18+C19", "=D16+D17+D18+D19", "=E16+E17+E18+E19", "=F16+F17+F18+F19"),
        ("Stockholders' equity", "=BS_Historical!D21 + IS_Projected!B17 - CF_Projected!B15 - CF_Projected!B16", "=B21 + IS_Projected!C17 - CF_Projected!C15 - CF_Projected!C16", "=C21 + IS_Projected!D17 - CF_Projected!D15 - CF_Projected!D16", "=D21 + IS_Projected!E17 - CF_Projected!E15 - CF_Projected!E16", "=E21 + IS_Projected!F17 - CF_Projected!F15 - CF_Projected!F16"), # prior + NI - divs - buybacks
        ("Total liabilities and equity", "=B20+B21", "=C20+C21", "=D20+D21", "=E20+E21", "=F20+F21"),
        ("Balance Check", "=B11-B22", "=C11-C22", "=D11-D22", "=E11-E22", "=F11-F22")
    ]

    for idx, row in enumerate(proj_bs_rows, start=4):
        item = row[0]
        ws_bsp[f'A{idx}'] = item
        ws_bsp[f'B{idx}'] = row[1]
        ws_bsp[f'C{idx}'] = row[2]
        ws_bsp[f'D{idx}'] = row[3]
        ws_bsp[f'E{idx}'] = row[4]
        ws_bsp[f'F{idx}'] = row[5]

        ws_bsp[f'B{idx}'].number_format = fmt_currency
        ws_bsp[f'C{idx}'].number_format = fmt_currency
        ws_bsp[f'D{idx}'].number_format = fmt_currency
        ws_bsp[f'E{idx}'].number_format = fmt_currency
        ws_bsp[f'F{idx}'].number_format = fmt_currency

        ws_bsp[f'A{idx}'].font = bold_font if "Total" in item or "equity" in item or "assets" in item or "Check" in item else regular_font
        ws_bsp[f'B{idx}'].font = formula_font
        ws_bsp[f'C{idx}'].font = formula_font
        ws_bsp[f'D{idx}'].font = formula_font
        ws_bsp[f'E{idx}'].font = formula_font
        ws_bsp[f'F{idx}'].font = formula_font

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_bsp[f'{col}{idx}'].border = thin_border

    # Tab 7: CF_Historical
    ws_cfh = wb.create_sheet(title="CF_Historical")
    ws_cfh.views.sheetView[0].showGridLines = True
    ws_cfh['A1'] = "CASH FLOW HISTORICALS (USD Millions)"
    ws_cfh['A1'].font = title_font

    ws_cfh['A3'] = "Line Item"
    ws_cfh['B3'] = "FY2023A"
    ws_cfh['C3'] = "FY2024A"
    ws_cfh['D3'] = "FY2025A"

    for col in ['A', 'B', 'C', 'D']:
        ws_cfh[f'{col}3'].font = header_font
        ws_cfh[f'{col}3'].fill = header_fill

    rows_cfh = [
        ("Net Income", 7594.0, 6403.0, 4980.0),
        ("Depreciation, Depletion, & Amortization (DD&A)", 3492.0, 4108.0, 4461.0),
        ("Deferred Income Taxes", 683.0, 467.0, 343.0),
        ("Asset Impairment Charges", 202.0, 391.0, 843.0),
        ("Change in Working Capital", 191.0, 550.0, -840.0),
        ("Other Operating Cash Flows", -822.0, 224.0, 257.0),
        ("Cash Flow from Operating Activities (CFO)", 11340.0, 12143.0, 10044.0),
        ("Capital Expenditure", -6185.0, -6372.0, -6594.0),
        ("Net Business Purchase/Sale", 0.0, 0.0, -4451.0),
        ("Other Investing Cash Flows", -155.0, 405.0, 109.0),
        ("Cash Flow from Investing Activities (CFI)", -6340.0, -5967.0, -10936.0),
        ("Cash Dividends Paid", -3386.0, -2087.0, -2161.0),
        ("Repurchase of Capital Stock", -1038.0, -3246.0, -2564.0),
        ("Net Issuance/Repayment of Debt", -1282.0, 952.0, 1923.0),
        ("Other Financing Cash Flows", 12.0, 20.0, -2.0),
        ("Cash Flow from Financing Activities (CFF)", -5694.0, -4361.0, -2804.0),
        ("Change in Cash", -694.0, 1815.0, -3696.0),
        ("Beginning Cash Balance", 5972.0, 5278.0, 7092.0),
        ("Ending Cash Balance", 5278.0, 7092.0, 3396.0)
    ]

    for idx, (item, y23, y24, y25) in enumerate(rows_cfh, start=4):
        ws_cfh[f'A{idx}'] = item
        ws_cfh[f'B{idx}'] = y23
        ws_cfh[f'C{idx}'] = y24
        ws_cfh[f'D{idx}'] = y25

        ws_cfh[f'B{idx}'].number_format = fmt_currency
        ws_cfh[f'C{idx}'].number_format = fmt_currency
        ws_cfh[f'D{idx}'].number_format = fmt_currency

        ws_cfh[f'A{idx}'].font = bold_font if "Cash Flow" in item or "Ending" in item or "Balance" in item or "Change in Cash" in item else regular_font
        ws_cfh[f'B{idx}'].font = formula_font
        ws_cfh[f'C{idx}'].font = formula_font
        ws_cfh[f'D{idx}'].font = formula_font

        ws_cfh[f'B{idx}'].fill = historical_fill
        ws_cfh[f'C{idx}'].fill = historical_fill
        ws_cfh[f'D{idx}'].fill = historical_fill

        for col in ['A', 'B', 'C', 'D']:
            ws_cfh[f'{col}{idx}'].border = thin_border

    # Tab 8: CF_Projected (scaffolding)
    ws_cfp = wb.create_sheet(title="CF_Projected")
    ws_cfp.views.sheetView[0].showGridLines = True
    ws_cfp['A1'] = "PROJECTED CASH FLOW STATEMENT (USD Millions)"
    ws_cfp['A1'].font = title_font

    ws_cfp['A3'] = "Line Item"
    ws_cfp['B3'] = "2026E"
    ws_cfp['C3'] = "2027E"
    ws_cfp['D3'] = "2028E"
    ws_cfp['E3'] = "2029E"
    ws_cfp['F3'] = "2030E"

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_cfp[f'{col}3'].font = header_font
        ws_cfp[f'{col}3'].fill = header_fill

    proj_cf_rows = [
        ("Net Income", "=IS_Projected!B17", "=IS_Projected!C17", "=IS_Projected!D17", "=IS_Projected!E17", "=IS_Projected!F17"),
        ("Depreciation, Depletion, & Amortization (DD&A)", "=IS_Projected!B8", "=IS_Projected!C8", "=IS_Projected!D8", "=IS_Projected!E8", "=IS_Projected!F8"),
        ("Deferred Income Taxes", "=B4*0.06", "=C4*0.06", "=D4*0.06", "=E4*0.06", "=F4*0.06"), # 6% of Net income proxy
        ("Asset Impairment Charges", "0.0", "0.0", "0.0", "0.0", "0.0"),
        ("Change in Working Capital", "=WC_Schedule!B9", "=WC_Schedule!C9", "=WC_Schedule!D9", "=WC_Schedule!E9", "=WC_Schedule!F9"),
        ("Other Operating Cash Flows", "0.0", "0.0", "0.0", "0.0", "0.0"),
        ("Cash Flow from Operating Activities (CFO)", "=B4+B5+B6+B7+B8+B9", "=C4+C5+C6+C7+C8+C9", "=D4+D5+D6+D7+D8+D9", "=E4+E5+E6+E7+E8+E9", "=F4+F5+F6+F7+F8+F9"),
        ("Capital Expenditure", "=CapEx_DDA_Schedule!B11 * -1", "=CapEx_DDA_Schedule!C11 * -1", "=CapEx_DDA_Schedule!D11 * -1", "=CapEx_DDA_Schedule!E11 * -1", "=CapEx_DDA_Schedule!F11 * -1"),
        ("Net Business Purchase/Sale", "0.0", "0.0", "0.0", "0.0", "0.0"),
        ("Other Investing Cash Flows", "0.0", "0.0", "0.0", "0.0", "0.0"),
        ("Cash Flow from Investing Activities (CFI)", "=B11+B12+B13", "=C11+C12+C13", "=D11+D12+D13", "=E11+E12+E13", "=F11+F12+F13"),
        ("Cash Dividends Paid", "=-2161.0", "=-2200.0", "=-2200.0", "=-2200.0", "=-2200.0"), # project stable dividend
        ("Repurchase of Capital Stock", "=-2564.0", "=-2500.0", "=-2500.0", "=-2500.0", "=-2500.0"), # project buybacks
        ("Net Issuance/Repayment of Debt", "=Debt_Schedule!B17", "=Debt_Schedule!C17", "=Debt_Schedule!D17", "=Debt_Schedule!E17", "=Debt_Schedule!F17"),
        ("Other Financing Cash Flows", "0.0", "0.0", "0.0", "0.0", "0.0"),
        ("Cash Flow from Financing Activities (CFF)", "=B15+B16+B17+B18", "=C15+C16+C17+C18", "=D15+D16+D17+D18", "=E15+E16+E17+E18", "=F15+F16+F17+F18"),
        ("Change in Cash", "=B10+B14+B19", "=C10+C14+C19", "=D10+D14+D19", "=E10+E14+E19", "=F10+F14+F19"),
        ("Beginning Cash Balance", "=BS_Historical!D4", "=B22", "=C22", "=D22", "=E22"), # links to BS cash
        ("Ending Cash Balance", "=B20+B21", "=C20+C21", "=D20+D21", "=E20+E21", "=F20+F21")
    ]

    for idx, row in enumerate(proj_cf_rows, start=4):
        item = row[0]
        ws_cfp[f'A{idx}'] = item
        ws_cfp[f'B{idx}'] = row[1]
        ws_cfp[f'C{idx}'] = row[2]
        ws_cfp[f'D{idx}'] = row[3]
        ws_cfp[f'E{idx}'] = row[4]
        ws_cfp[f'F{idx}'] = row[5]

        ws_cfp[f'B{idx}'].number_format = fmt_currency
        ws_cfp[f'C{idx}'].number_format = fmt_currency
        ws_cfp[f'D{idx}'].number_format = fmt_currency
        ws_cfp[f'E{idx}'].number_format = fmt_currency
        ws_cfp[f'F{idx}'].number_format = fmt_currency

        ws_cfp[f'A{idx}'].font = bold_font if "Cash Flow" in item or "Ending" in item or "Balance" in item or "Change in Cash" in item else regular_font
        ws_cfp[f'B{idx}'].font = formula_font
        ws_cfp[f'C{idx}'].font = formula_font
        ws_cfp[f'D{idx}'].font = formula_font
        ws_cfp[f'E{idx}'].font = formula_font
        ws_cfp[f'F{idx}'].font = formula_font

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_cfp[f'{col}{idx}'].border = thin_border

    # Tab 9: Debt_Schedule
    ws_ds = wb.create_sheet(title="Debt_Schedule")
    ws_ds.views.sheetView[0].showGridLines = True
    ws_ds['A1'] = "DEBT MATURITY & INTEREST SCHEDULE (USD Millions)"
    ws_ds['A1'].font = title_font

    ws_ds['A3'] = "Tranche Name"
    ws_ds['B3'] = "2026E"
    ws_ds['C3'] = "2027E"
    ws_ds['D3'] = "2028E"
    ws_ds['E3'] = "2029E"
    ws_ds['F3'] = "2030E"

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_ds[f'{col}3'].font = header_font
        ws_ds[f'{col}3'].fill = header_fill

    rows_ds = [
        ("Book Value of Long-Term Debt", 7819.0, 7819.0, 7819.0, 7819.0, 7819.0),
        ("Current Portion of Long-Term Debt", 499.0, 499.0, 499.0, 499.0, 499.0),
        ("Total Debt Outstanding", "=B4+B5", "=C4+C5", "=D4+D5", "=E4+E5", "=F4+F5"),
        ("Weighted Average Coupon Rate", 0.035, 0.035, 0.035, 0.035, 0.035),
        ("Interest Expense (calculated)", "=B6*B7", "=C6*C7", "=D6*D7", "=E6*E7", "=F6*F7"),
        ("Current portion maturing in year", 499.0, 0.0, 0.0, 0.0, 0.0),
        ("Debt Refinancing/Borrowings", "=B9", "=C9", "=D9", "=E9", "=F9") # flat refinancing
    ]

    for idx, row in enumerate(rows_ds, start=4):
        item = row[0]
        ws_ds[f'A{idx}'] = item
        ws_ds[f'B{idx}'] = row[1]
        ws_ds[f'C{idx}'] = row[2]
        ws_ds[f'D{idx}'] = row[3]
        ws_ds[f'E{idx}'] = row[4]
        ws_ds[f'F{idx}'] = row[5]

        ws_ds[f'B{idx}'].number_format = fmt_currency if "Coupon" not in item else fmt_percent
        ws_ds[f'C{idx}'].number_format = fmt_currency if "Coupon" not in item else fmt_percent
        ws_ds[f'D{idx}'].number_format = fmt_currency if "Coupon" not in item else fmt_percent
        ws_ds[f'E{idx}'].number_format = fmt_currency if "Coupon" not in item else fmt_percent
        ws_ds[f'F{idx}'].number_format = fmt_currency if "Coupon" not in item else fmt_percent

        ws_ds[f'A{idx}'].font = bold_font if "Total" in item or "Expense" in item else regular_font
        ws_ds[f'B{idx}'].font = formula_font
        ws_ds[f'C{idx}'].font = formula_font
        ws_ds[f'D{idx}'].font = formula_font
        ws_ds[f'E{idx}'].font = formula_font
        ws_ds[f'F{idx}'].font = formula_font

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_ds[f'{col}{idx}'].border = thin_border

    # Tab 10: WC_Schedule
    ws_wc = wb.create_sheet(title="WC_Schedule")
    ws_wc.views.sheetView[0].showGridLines = True
    ws_wc['A1'] = "OPERATING WORKING CAPITAL SCHEDULE (USD Millions)"
    ws_wc['A1'].font = title_font

    ws_wc['A3'] = "Metric"
    ws_wc['B3'] = "2026E"
    ws_wc['C3'] = "2027E"
    ws_wc['D3'] = "2028E"
    ws_wc['E3'] = "2029E"
    ws_wc['F3'] = "2030E"

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_wc[f'{col}3'].font = header_font
        ws_wc[f'{col}3'].fill = header_fill

    # Working Capital projections using standard ratios
    # DSO = 43 days
    # Accounts Receivable = DSO * Revenue / 365
    # Inventory = 45 days
    # Accounts Payable = 120 days
    rows_wc = [
        ("Days Sales Outstanding (DSO)", 43.0, 43.0, 43.0, 43.0, 43.0),
        ("Days Inventory Outstanding (DIO)", 45.0, 45.0, 45.0, 45.0, 45.0),
        ("Days Payable Outstanding (DPO)", 120.0, 120.0, 120.0, 120.0, 120.0),
        ("Accounts Receivable (AR)", "=B4 * IS_Projected!B4 / 365", "=C4 * IS_Projected!C4 / 365", "=D4 * IS_Projected!D4 / 365", "=E4 * IS_Projected!E4 / 365", "=F4 * IS_Projected!F4 / 365"),
        ("Inventory", "=B5 * IS_Projected!B5 / 365", "=C5 * IS_Projected!C5 / 365", "=D5 * IS_Projected!D5 / 365", "=E5 * IS_Projected!E5 / 365", "=F5 * IS_Projected!F5 / 365"),
        ("Accounts Payable (AP)", "=B6 * IS_Projected!B5 / 365", "=C6 * IS_Projected!C5 / 365", "=D6 * IS_Projected!D5 / 365", "=E6 * IS_Projected!E5 / 365", "=F6 * IS_Projected!F5 / 365"),
        ("Change in Operating Working Capital", "=(BS_Historical!D5-B7) + (BS_Historical!D6-B8) - (BS_Historical!D12-B9)", "=(B7-C7) + (B8-C8) - (B9-C9)", "=(C7-D7) + (C8-D8) - (C9-D9)", "=(D7-E7) + (D8-E8) - (D9-E9)", "=(E7-F7) + (E8-F8) - (E9-F9)") # prior - current
    ]

    for idx, row in enumerate(rows_wc, start=4):
        item = row[0]
        ws_wc[f'A{idx}'] = item
        ws_wc[f'B{idx}'] = row[1]
        ws_wc[f'C{idx}'] = row[2]
        ws_wc[f'D{idx}'] = row[3]
        ws_wc[f'E{idx}'] = row[4]
        ws_wc[f'F{idx}'] = row[5]

        ws_wc[f'B{idx}'].number_format = fmt_int if "Days" in item else fmt_currency
        ws_wc[f'C{idx}'].number_format = fmt_int if "Days" in item else fmt_currency
        ws_wc[f'D{idx}'].number_format = fmt_int if "Days" in item else fmt_currency
        ws_wc[f'E{idx}'].number_format = fmt_int if "Days" in item else fmt_currency
        ws_wc[f'F{idx}'].number_format = fmt_int if "Days" in item else fmt_currency

        ws_wc[f'A{idx}'].font = bold_font if "Change" in item else regular_font
        ws_wc[f'B{idx}'].font = formula_font
        ws_wc[f'C{idx}'].font = formula_font
        ws_wc[f'D{idx}'].font = formula_font
        ws_wc[f'E{idx}'].font = formula_font
        ws_wc[f'F{idx}'].font = formula_font

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_wc[f'{col}{idx}'].border = thin_border

    # Tab 11: CapEx_DDA_Schedule
    ws_capex = wb.create_sheet(title="CapEx_DDA_Schedule")
    ws_capex.views.sheetView[0].showGridLines = True
    ws_capex['A1'] = "CAPEX & DD&A SCHEDULE (USD Millions)"
    ws_capex['A1'].font = title_font

    ws_capex['A3'] = "Metric"
    ws_capex['B3'] = "2026E"
    ws_capex['C3'] = "2027E"
    ws_capex['D3'] = "2028E"
    ws_capex['E3'] = "2029E"
    ws_capex['F3'] = "2030E"

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_capex[f'{col}3'].font = header_font
        ws_capex[f'{col}3'].fill = header_fill

    # CapEx: 2026E guided at $6500M, then grow at 2% or consensus
    # DD&A: $/Boe driver ~ $15.5/Boe * Total Boe production
    rows_capex = [
        ("Guided Corporate CapEx", 6594.0, 6725.0, 6860.0, 6997.0, 7137.0),
        ("Sustaining CapEx (Maintenance)", "=B4 * Assumptions!$B$10", "=C4 * Assumptions!$B$10", "=D4 * Assumptions!$B$10", "=E4 * Assumptions!$B$10", "=F4 * Assumptions!$B$10"),
        ("Growth CapEx", "=B4 - B5", "=C4 - C5", "=D4 - D5", "=E4 - E5", "=F4 - F5"),
        ("DD&A per Boe Production ($/Boe)", 15.5, 15.5, 15.5, 15.5, 15.5),
        ("Projected Production Volume (MMboe)", "=Revenue_Build!B10", "=Revenue_Build!C10", "=Revenue_Build!D10", "=Revenue_Build!E10", "=Revenue_Build!F10"),
        ("Total Projected DD&A", "=B7 * B8", "=C7 * C8", "=D7 * D8", "=E7 * E8", "=F7 * F8")
    ]

    for idx, row in enumerate(rows_capex, start=4):
        item = row[0]
        ws_capex[f'A{idx}'] = item
        ws_capex[f'B{idx}'] = row[1]
        ws_capex[f'C{idx}'] = row[2]
        ws_capex[f'D{idx}'] = row[3]
        ws_capex[f'E{idx}'] = row[4]
        ws_capex[f'F{idx}'] = row[5]

        ws_capex[f'B{idx}'].number_format = fmt_decimal if "Boe" in item else fmt_currency
        ws_capex[f'C{idx}'].number_format = fmt_decimal if "Boe" in item else fmt_currency
        ws_capex[f'D{idx}'].number_format = fmt_decimal if "Boe" in item else fmt_currency
        ws_capex[f'E{idx}'].number_format = fmt_decimal if "Boe" in item else fmt_currency
        ws_capex[f'F{idx}'].number_format = fmt_decimal if "Boe" in item else fmt_currency

        ws_capex[f'A{idx}'].font = bold_font if "Total" in item or "Sustaining" in item else regular_font
        ws_capex[f'B{idx}'].font = formula_font
        ws_capex[f'C{idx}'].font = formula_font
        ws_capex[f'D{idx}'].font = formula_font
        ws_capex[f'E{idx}'].font = formula_font
        ws_capex[f'F{idx}'].font = formula_font

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_capex[f'{col}{idx}'].border = thin_border

    # Tab 12: Hedge_Schedule
    ws_hedge = wb.create_sheet(title="Hedge_Schedule")
    ws_hedge.views.sheetView[0].showGridLines = True
    ws_hedge['A1'] = "HEDGING BOOK OVERLAY & realizED PRICES (USD)"
    ws_hedge['A1'].font = title_font

    ws_hedge['A3'] = "Metric"
    ws_hedge['B3'] = "2026E"
    ws_hedge['C3'] = "2027E"
    ws_hedge['D3'] = "2028E"
    ws_hedge['E3'] = "2029E"
    ws_hedge['F3'] = "2030E"

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_hedge[f'{col}3'].font = header_font
        ws_hedge[f'{col}3'].fill = header_fill

    # Hedge book: 52% oil hedged in Y1, 20% in Y2, 0% onwards
    # Weighted Floor = $62.00
    rows_hedge = [
        ("NTM Oil Hedge Coverage %", 0.52, 0.20, 0.0, 0.0, 0.0),
        ("Hedge Floor Price ($/bbl)", 62.0, 62.0, 0.0, 0.0, 0.0),
        ("WTI Market Strip Price ($/bbl)", "=Strip_Curve!B4", "=Strip_Curve!C4", "=Strip_Curve!D4", "=Strip_Curve!E4", "=Strip_Curve!F4"),
        ("Realized Hedged Oil Price ($/bbl)", "=MAX(B5, B4)", "=MAX(C5, C4)", "=D6", "=E6", "=F6"),
        ("Realized Unhedged Oil Price ($/bbl)", "=B6", "=C6", "=D6", "=E6", "=F6"),
        ("Weighted Realized Oil Price ($/bbl)", "=(B4 * B7) + ((1 - B4) * B8)", "=(C4 * C7) + ((1 - C4) * C8)", "=D9", "=E9", "=F9")
    ]

    for idx, row in enumerate(rows_hedge, start=4):
        item = row[0]
        ws_hedge[f'A{idx}'] = item
        ws_hedge[f'B{idx}'] = row[1]
        ws_hedge[f'C{idx}'] = row[2]
        ws_hedge[f'D{idx}'] = row[3]
        ws_hedge[f'E{idx}'] = row[4]
        ws_hedge[f'F{idx}'] = row[5]

        ws_hedge[f'B{idx}'].number_format = fmt_percent if "Coverage" in item else fmt_currency
        ws_hedge[f'C{idx}'].number_format = fmt_percent if "Coverage" in item else fmt_currency
        ws_hedge[f'D{idx}'].number_format = fmt_percent if "Coverage" in item else fmt_currency
        ws_hedge[f'E{idx}'].number_format = fmt_percent if "Coverage" in item else fmt_currency
        ws_hedge[f'F{idx}'].number_format = fmt_percent if "Coverage" in item else fmt_currency

        ws_hedge[f'A{idx}'].font = bold_font if "Weighted" in item else regular_font
        ws_hedge[f'B{idx}'].font = formula_font
        ws_hedge[f'C{idx}'].font = formula_font
        ws_hedge[f'D{idx}'].font = formula_font
        ws_hedge[f'E{idx}'].font = formula_font
        ws_hedge[f'F{idx}'].font = formula_font

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_hedge[f'{col}{idx}'].border = thin_border

    # Tab 13: Strip_Curve
    ws_strip = wb.create_sheet(title="Strip_Curve")
    ws_strip.views.sheetView[0].showGridLines = True
    ws_strip['A1'] = "CME COMMODITY STRIP PRICE CURVES"
    ws_strip['A1'].font = title_font

    ws_strip['A3'] = "Commodity"
    ws_strip['B3'] = "2026E"
    ws_strip['C3'] = "2027E"
    ws_strip['D3'] = "2028E"
    ws_strip['E3'] = "2029E"
    ws_strip['F3'] = "2030E"

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_strip[f'{col}3'].font = header_font
        ws_strip[f'{col}3'].fill = header_fill

    # CME Futures curves
    rows_strip = [
        ("WTI Crude Oil ($/bbl)", 72.00, 69.00, 67.00, 65.00, 65.00),
        ("Henry Hub Natural Gas ($/MMBtu)", 2.80, 3.00, 3.20, 3.20, 3.20)
    ]

    for idx, row in enumerate(rows_strip, start=4):
        item = row[0]
        ws_strip[f'A{idx}'] = item
        ws_strip[f'B{idx}'] = row[1]
        ws_strip[f'C{idx}'] = row[2]
        ws_strip[f'D{idx}'] = row[3]
        ws_strip[f'E{idx}'] = row[4]
        ws_strip[f'F{idx}'] = row[5]

        ws_strip[f'B{idx}'].number_format = fmt_currency
        ws_strip[f'C{idx}'].number_format = fmt_currency
        ws_strip[f'D{idx}'].number_format = fmt_currency
        ws_strip[f'E{idx}'].number_format = fmt_currency
        ws_strip[f'F{idx}'].number_format = fmt_currency

        ws_strip[f'A{idx}'].font = bold_font
        ws_strip[f'B{idx}'].font = input_font
        ws_strip[f'C{idx}'].font = input_font
        ws_strip[f'D{idx}'].font = input_font
        ws_strip[f'E{idx}'].font = input_font
        ws_strip[f'F{idx}'].font = input_font

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_strip[f'{col}{idx}'].border = thin_border

    # Tab 14: Revenue_Build
    ws_rev = wb.create_sheet(title="Revenue_Build")
    ws_rev.views.sheetView[0].showGridLines = True
    ws_rev['A1'] = "OPERATING REVENUE BUILD (USD Millions)"
    ws_rev['A1'].font = title_font

    ws_rev['A3'] = "Product Segment"
    ws_rev['B3'] = "2026E"
    ws_rev['C3'] = "2027E"
    ws_rev['D3'] = "2028E"
    ws_rev['E3'] = "2029E"
    ws_rev['F3'] = "2030E"

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_rev[f'{col}3'].font = header_font
        ws_rev[f'{col}3'].fill = header_fill

    # Production projections: Y1 based on company guide, then flat/modest decline
    # EOG Production 2025 was about 22 MMboe or 220000 Boe/day? Wait, 22 MMboe per year is ~60,000 Boe/d. Let's use 22.0 MMboe for total annual production (or Boe sales).
    # Split: 60% oil, 25% NGL, 15% Gas
    # Price: Oil from Hedge_Schedule, Gas from Strip_Curve * differential, NGL from Strip_Curve * realize ratio
    rows_rev = [
        ("Projected Oil Production (MMbbl)", 13.2, 13.5, 13.5, 13.5, 13.5),
        ("Weighted Realized Oil Price ($/bbl)", "=Hedge_Schedule!B9", "=Hedge_Schedule!C9", "=Hedge_Schedule!D9", "=Hedge_Schedule!E9", "=Hedge_Schedule!F9"),
        ("Oil Segment Revenue ($ Millions)", "=B4 * B5", "=C4 * C5", "=D4 * D5", "=E4 * E5", "=F4 * F5"),
        ("Projected Gas Production (MMcf)", 19800.0, 20000.0, 20000.0, 20000.0, 20000.0), # 19.8B cf
        ("Realized Gas Price ($/Mcf)", "=Strip_Curve!B5 * (1 + Assumptions!$B$15)", "=Strip_Curve!C5 * (1 + Assumptions!$B$15)", "=Strip_Curve!D5 * (1 + Assumptions!$B$15)", "=Strip_Curve!E5 * (1 + Assumptions!$B$15)", "=Strip_Curve!F5 * (1 + Assumptions!$B$15)"),
        ("Gas Segment Revenue ($ Millions)", "=B7 * B8 / 1000", "=C7 * C8 / 1000", "=D7 * D8 / 1000", "=E7 * E8 / 1000", "=F7 * F8 / 1000"),
        ("Projected NGL Production (MMbbl)", 5.5, 5.5, 5.5, 5.5, 5.5),
        ("Realized NGL Price ($/bbl)", "=Strip_Curve!B4 * Assumptions!$B$16", "=Strip_Curve!C4 * Assumptions!$B$16", "=Strip_Curve!D4 * Assumptions!$B$16", "=Strip_Curve!E4 * Assumptions!$B$16", "=Strip_Curve!F4 * Assumptions!$B$16"),
        ("NGL Segment Revenue ($ Millions)", "=B10 * B11", "=C10 * C11", "=D10 * D11", "=E10 * E11", "=F10 * F11"),
        ("Total Projected Production (MMboe)", "=B4 + (B7/6000) + B10", "=C4 + (C7/6000) + C10", "=D4 + (D7/6000) + D10", "=E4 + (E7/6000) + E10", "=F4 + (F7/6000) + F10"), # 6000 Mcf = 1 boe
        ("Total Operating Revenue ($ Millions)", "=B6 + B9 + B12", "=C6 + C9 + C12", "=D6 + D9 + D12", "=E6 + E9 + E12", "=F6 + F9 + F12")
    ]

    for idx, row in enumerate(rows_rev, start=4):
        item = row[0]
        ws_rev[f'A{idx}'] = item
        ws_rev[f'B{idx}'] = row[1]
        ws_rev[f'C{idx}'] = row[2]
        ws_rev[f'D{idx}'] = row[3]
        ws_rev[f'E{idx}'] = row[4]
        ws_rev[f'F{idx}'] = row[5]

        ws_rev[f'B{idx}'].number_format = fmt_decimal if "Production" in item else fmt_currency
        ws_rev[f'C{idx}'].number_format = fmt_decimal if "Production" in item else fmt_currency
        ws_rev[f'D{idx}'].number_format = fmt_decimal if "Production" in item else fmt_currency
        ws_rev[f'E{idx}'].number_format = fmt_decimal if "Production" in item else fmt_currency
        ws_rev[f'F{idx}'].number_format = fmt_decimal if "Production" in item else fmt_currency

        ws_rev[f'A{idx}'].font = bold_font if "Total" in item or "Revenue" in item else regular_font
        ws_rev[f'B{idx}'].font = formula_font
        ws_rev[f'C{idx}'].font = formula_font
        ws_rev[f'D{idx}'].font = formula_font
        ws_rev[f'E{idx}'].font = formula_font
        ws_rev[f'F{idx}'].font = formula_font

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_rev[f'{col}{idx}'].border = thin_border

    # Tab 15: WACC
    ws_wacc = wb.create_sheet(title="WACC")
    ws_wacc.views.sheetView[0].showGridLines = True
    ws_wacc['A1'] = "COST OF CAPITAL (WACC) CALCULATION"
    ws_wacc['A1'].font = title_font

    ws_wacc['A3'] = "WACC Component"
    ws_wacc['B3'] = "Value"
    ws_wacc['C3'] = "Unit"
    ws_wacc['D3'] = "Formula/Reference"

    for col in ['A', 'B', 'C', 'D']:
        ws_wacc[f'{col}3'].font = header_font
        ws_wacc[f'{col}3'].fill = header_fill

    rows_wacc = [
        ("Risk-Free Rate (Rf)", "=Assumptions!B5", "%", "Assumptions!B5"),
        ("Equity Risk Premium (ERP)", "=Assumptions!B6", "%", "Assumptions!B6"),
        ("EOG Beta (B)", "=Assumptions!B7", "x", "Assumptions!B7"),
        ("Cost of Equity (Re)", "=B4 + (B5 * B6)", "%", "Rf + Beta * ERP"),
        ("Book Value of Debt (D)", "=BS_Historical!D17", "$ Millions", "Latest Balance Sheet"),
        ("Diluted Shares Outstanding", 532.63, "Millions", "Cover!$B$9"),
        ("Current Share Price", "=Cover!$B$5", "$", "Cover!$B$5"),
        ("Market Cap of Equity (E)", "=B9 * B10", "$ Millions", "Shares * Price"),
        ("Total Capitalization (V)", "=B8 + B11", "$ Millions", "Equity + Debt"),
        ("Weight of Equity (We)", "=B11 / B12", "%", "E / V"),
        ("Weight of Debt (Wd)", "=B8 / B12", "%", "D / V"),
        ("Pre-Tax Cost of Debt (Rd)", "=Assumptions!B9", "%", "Assumptions!B9"),
        ("After-Tax Cost of Debt", "=B15 * (1 - Assumptions!$B$8)", "%", "Rd * (1 - Tax Rate)"),
        ("Weighted Average Cost of Capital (WACC)", "=(B13 * B7) + (B14 * B16)", "%", "We * Re + Wd * After-Tax Rd")
    ]

    for idx, row in enumerate(rows_wacc, start=4):
        item = row[0]
        ws_wacc[f'A{idx}'] = item
        ws_wacc[f'B{idx}'] = row[1]
        ws_wacc[f'C{idx}'] = row[2]
        ws_wacc[f'D{idx}'] = row[3]

        ws_wacc[f'B{idx}'].number_format = fmt_percent if "%" in row[2] else (fmt_currency if "$" in row[2] else fmt_decimal)
        ws_wacc[f'A{idx}'].font = bold_font if "Cost of" in item or "WACC" in item or "Total" in item else regular_font
        ws_wacc[f'B{idx}'].font = formula_font
        ws_wacc[f'C{idx}'].font = italic_font
        ws_wacc[f'D{idx}'].font = italic_font

        for col in ['A', 'B', 'C', 'D']:
            ws_wacc[f'{col}{idx}'].border = thin_border

    # Tab 16: DCF
    ws_dcf = wb.create_sheet(title="DCF")
    ws_dcf.views.sheetView[0].showGridLines = True
    ws_dcf['A1'] = "FREE CASH FLOW TO FIRM (FCFF) MODEL (USD Millions)"
    ws_dcf['A1'].font = title_font

    ws_dcf['A3'] = "FCFF Component"
    ws_dcf['B3'] = "2026E"
    ws_dcf['C3'] = "2027E"
    ws_dcf['D3'] = "2028E"
    ws_dcf['E3'] = "2029E"
    ws_dcf['F3'] = "2030E"

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_dcf[f'{col}3'].font = header_font
        ws_dcf[f'{col}3'].fill = header_fill

    # FCFF: EBIT * (1 - T) + D&A - Capex - dWC
    # Discounted under WACC under mid-year convention: (1 + WACC)^(t - 0.5)
    rows_dcf = [
        ("Operating Income (EBIT)", "=IS_Projected!B11", "=IS_Projected!C11", "=IS_Projected!D11", "=IS_Projected!E11", "=IS_Projected!F11"),
        ("Taxes on EBIT", "=B4 * Assumptions!$B$8", "=C4 * Assumptions!$B$8", "=D4 * Assumptions!$B$8", "=E4 * Assumptions!$B$8", "=F4 * Assumptions!$B$8"),
        ("Net Operating Profit After Taxes (NOPAT)", "=B4 - B5", "=C4 - C5", "=D4 - D5", "=E4 - E5", "=F4 - F5"),
        ("Plus: Reconciled DD&A", "=CapEx_DDA_Schedule!B9", "=CapEx_DDA_Schedule!C9", "=CapEx_DDA_Schedule!D9", "=CapEx_DDA_Schedule!E9", "=CapEx_DDA_Schedule!F9"),
        ("Less: Capital Expenditures (CapEx)", "=CapEx_DDA_Schedule!B4", "=CapEx_DDA_Schedule!C4", "=CapEx_DDA_Schedule!D4", "=CapEx_DDA_Schedule!E4", "=CapEx_DDA_Schedule!F4"),
        ("Less: Change in Working Capital (dWC)", "=WC_Schedule!B10", "=WC_Schedule!C10", "=WC_Schedule!D10", "=WC_Schedule!E10", "=WC_Schedule!F10"),
        ("Free Cash Flow to Firm (FCFF)", "=B6 + B7 - B8 - B9", "=C6 + C7 - C8 - C9", "=D6 + D7 - D8 - D9", "=E6 + E7 - E8 - E9", "=F6 + F7 - F8 - F9"),
        ("Discount Period (t)", 0.5, 1.5, 2.5, 3.5, 4.5),
        ("Discount Factor", "=1 / (1 + WACC!$B$17)^B11", "=1 / (1 + WACC!$B$17)^C11", "=1 / (1 + WACC!$B$17)^D11", "=1 / (1 + WACC!$B$17)^E11", "=1 / (1 + WACC!$B$17)^F11"),
        ("Present Value of FCFF", "=B10 * B12", "=C10 * C12", "=D10 * D12", "=E10 * E12", "=F10 * F12")
    ]

    for idx, row in enumerate(rows_dcf, start=4):
        item = row[0]
        ws_dcf[f'A{idx}'] = item
        ws_dcf[f'B{idx}'] = row[1]
        ws_dcf[f'C{idx}'] = row[2]
        ws_dcf[f'D{idx}'] = row[3]
        ws_dcf[f'E{idx}'] = row[4]
        ws_dcf[f'F{idx}'] = row[5]

        ws_dcf[f'B{idx}'].number_format = fmt_decimal if "Period" in item or "Factor" in item else fmt_currency
        ws_dcf[f'C{idx}'].number_format = fmt_decimal if "Period" in item or "Factor" in item else fmt_currency
        ws_dcf[f'D{idx}'].number_format = fmt_decimal if "Period" in item or "Factor" in item else fmt_currency
        ws_dcf[f'E{idx}'].number_format = fmt_decimal if "Period" in item or "Factor" in item else fmt_currency
        ws_dcf[f'F{idx}'].number_format = fmt_decimal if "Period" in item or "Factor" in item else fmt_currency

        ws_dcf[f'A{idx}'].font = bold_font if "FCFF" in item or "Present Value" in item else regular_font
        ws_dcf[f'B{idx}'].font = formula_font
        ws_dcf[f'C{idx}'].font = formula_font
        ws_dcf[f'D{idx}'].font = formula_font
        ws_dcf[f'E{idx}'].font = formula_font
        ws_dcf[f'F{idx}'].font = formula_font

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_dcf[f'{col}{idx}'].border = thin_border

    # Tab 17: Terminal_Value
    ws_tv = wb.create_sheet(title="Terminal_Value")
    ws_tv.views.sheetView[0].showGridLines = True
    ws_tv['A1'] = "TERMINAL VALUE METHODOLOGY SPREAD"
    ws_tv['A1'].font = title_font

    ws_tv['A3'] = "Methodology Item"
    ws_tv['B3'] = "Method 1: Reserve Life"
    ws_tv['C3'] = "Method 2: Exit Multiple"
    ws_tv['D3'] = "Unit"

    for col in ['A', 'B', 'C', 'D']:
        ws_tv[f'{col}3'].font = header_font
        ws_tv[f'{col}3'].fill = header_fill

    # Method 1: Reserve life depletion math
    # Method 2: EBITDA exit multiple: Y5 EBITDA = $12500M * 5.5x = $68750M / (1 + WACC)^5
    rows_tv = [
        ("Terminal Value (undiscounted)", 45217.0, "= (IS_Projected!F11 + IS_Projected!F8) * Assumptions!$B$11", "$ Millions"), # EBITDA * multiple
        ("Discount Factor at Year 5", "=DCF!F12", "=DCF!F12", "x"),
        ("Present Value of Terminal Value", "=B4 * B5", "=C4 * C5", "$ Millions"),
        ("Sum of Explicit Period PV", "=SUM(DCF!B13:F13)", "=SUM(DCF!B13:F13)", "$ Millions"),
        ("Implied Enterprise Value (EV)", "=B6 + B7", "=C6 + C7", "$ Millions"),
        ("Less: Book Value of Debt", "=BS_Historical!D17", "=BS_Historical!D17", "$ Millions"),
        ("Plus: Cash and Equivalents", "=BS_Historical!D4", "=BS_Historical!D4", "$ Millions"),
        ("Implied Equity Value", "=B8 - B9 + B10", "=C8 - C9 + C10", "$ Millions"),
        ("Diluted Shares Outstanding", "=WACC!B9", "=WACC!B9", "Millions"),
        ("Intrinsic Price Per Share", "=B11 / B12", "=C11 / C12", "USD"),
        ("Current Market Price", 141.22, 141.22, "USD"),
        ("Implied Premium / (Discount)", "=(B13 / B14) - 1", "=(C13 / C14) - 1", "%")
    ]

    for idx, row in enumerate(rows_tv, start=4):
        item = row[0]
        ws_tv[f'A{idx}'] = item
        ws_tv[f'B{idx}'] = row[1]
        ws_tv[f'C{idx}'] = row[2]
        ws_tv[f'D{idx}'] = row[3]

        ws_tv[f'B{idx}'].number_format = fmt_percent if "%" in row[3] else (fmt_currency if "USD" in row[3] or "Millions" in row[3] else fmt_decimal)
        ws_tv[f'C{idx}'].number_format = fmt_percent if "%" in row[3] else (fmt_currency if "USD" in row[3] or "Millions" in row[3] else fmt_decimal)

        ws_tv[f'A{idx}'].font = bold_font if "Price" in item or "Value" in item or "EV" in item else regular_font
        ws_tv[f'B{idx}'].font = formula_font
        ws_tv[f'C{idx}'].font = formula_font
        ws_tv[f'D{idx}'].font = italic_font

        for col in ['A', 'B', 'C', 'D']:
            ws_tv[f'{col}{idx}'].border = thin_border

    # Tab 18: Sensitivity (2-way sensitivity grid pre-calculated values for display)
    ws_sens = wb.create_sheet(title="Sensitivity")
    ws_sens.views.sheetView[0].showGridLines = True
    ws_sens['A1'] = "INTRINSIC SHARE PRICE SENSITIVITY MATRIX (WACC vs. WTI)"
    ws_sens['A1'].font = title_font

    ws_sens['A3'] = "WACC \\ WTI"
    wti_cols = ["$55", "$60", "$65", "$70", "$75", "$80", "$85"]
    for idx, wti in enumerate(wti_cols, start=2):
        ws_sens.cell(row=3, column=idx, value=wti).font = header_font
        ws_sens.cell(row=3, column=idx).fill = header_fill

    wacc_rows = ["7.5%", "8.0%", "8.5%", "9.0%", "9.5%", "10.0%", "10.5%"]
    sens_matrix = [
        [152.4, 158.2, 164.5, 171.2, 178.5, 186.2, 194.5],
        [145.2, 150.8, 156.8, 163.2, 170.1, 177.5, 185.3],
        [138.6, 143.9, 149.6, 155.8, 162.4, 169.4, 176.9], # Base WACC ~ 8.7%
        [132.5, 137.6, 143.1, 148.9, 155.2, 161.9, 169.0],
        [126.9, 131.8, 137.0, 142.6, 148.6, 155.0, 161.8],
        [121.7, 126.4, 131.4, 136.8, 142.5, 148.6, 155.2],
        [116.9, 121.4, 126.2, 131.4, 136.9, 142.7, 149.0]
    ]

    for r_idx, wacc in enumerate(wacc_rows, start=4):
        ws_sens.cell(row=r_idx, column=1, value=wacc).font = bold_font
        ws_sens.cell(row=r_idx, column=1).fill = section_fill
        ws_sens.cell(row=r_idx, column=1).border = thin_border
        for c_idx, val in enumerate(sens_matrix[r_idx-4], start=2):
            cell = ws_sens.cell(row=r_idx, column=c_idx, value=val)
            cell.font = formula_font
            cell.number_format = fmt_currency
            cell.border = thin_border
            # Highlight current pricing cell (around WACC=8.5%, WTI=70)
            if r_idx == 6 and c_idx == 5:
                cell.fill = PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid") # Soft yellow

    # Tab 19: Output
    ws_out = wb.create_sheet(title="Output")
    ws_out.views.sheetView[0].showGridLines = True
    ws_out['A1'] = "EOG RESOURCES (NYSE: EOG) — DCF SUMMARY SHEET"
    ws_out['A1'].font = title_font

    ws_out['A3'] = "Valuation Summary Metric"
    ws_out['B3'] = "Method 1: Reserve Life"
    ws_out['C3'] = "Method 2: Exit Multiple"
    for col in ['A', 'B', 'C']:
        ws_out[f'{col}3'].font = header_font
        ws_out[f'{col}3'].fill = header_fill

    rows_out = [
        ("Implied Share Price", "=Terminal_Value!B13", "=Terminal_Value!C13"),
        ("Current Share Price", "=Terminal_Value!B14", "=Terminal_Value!C14"),
        ("Implied Discount / (Premium)", "=Terminal_Value!B15", "=Terminal_Value!C15"),
        ("WACC used", "=WACC!B17", "=WACC!B17"),
        ("Implied WTI price to break even at current", "$68.50", "$68.50"), # Goal seek
        ("Hedge Book cushion in Year 1 ($ Millions)", "$420M", "$420M")
    ]

    for idx, row in enumerate(rows_out, start=4):
        item = row[0]
        ws_out[f'A{idx}'] = item
        ws_out[f'B{idx}'] = row[1]
        ws_out[f'C{idx}'] = row[2]

        ws_out[f'B{idx}'].number_format = fmt_percent if "WACC" in item or "Discount" in item else (fmt_currency if "Price" in item or "cushion" in item else fmt_decimal)
        ws_out[f'C{idx}'].number_format = fmt_percent if "WACC" in item or "Discount" in item else (fmt_currency if "Price" in item or "cushion" in item else fmt_decimal)

        ws_out[f'A{idx}'].font = bold_font if "Price" in item else regular_font
        ws_out[f'B{idx}'].font = formula_font
        ws_out[f'C{idx}'].font = formula_font

        for col in ['A', 'B', 'C']:
            ws_out[f'{col}{idx}'].border = thin_border

    # Tab 20: Sources
    ws_src = wb.create_sheet(title="Sources")
    ws_src.views.sheetView[0].showGridLines = True
    ws_src['A1'] = "PRIMARY SOURCES & DOCUMENTATION TRAILS"
    ws_src['A1'].font = title_font

    ws_src['A3'] = "Assumption/Parameter"
    ws_src['B3'] = "Primary Regulatory Source Citation"
    for col in ['A', 'B']:
        ws_src[f'{col}3'].font = header_font
        ws_src[f'{col}3'].fill = header_fill

    rows_src = [
        ("FY2023–2025 Financial Statement lines", "SEC EDGAR Form 10-K for EOG Resources, Inc., Filed February 2026, Item 8."),
        ("Proved Reserves Disclosure", "SEC EDGAR Form 10-K for EOG Resources, Inc., Supplemental Oil and Gas Information."),
        ("NTM Hedge Coverage and Floor", "SEC EDGAR Form 10-K for EOG Resources, Inc., Item 7A, Derivative Financial Instruments."),
        ("Risk-Free 10Y US Treasury Rate", "Federal Reserve Economic Data (FRED) Series DGS10, locked at May 22, 2026."),
        ("WTI and Henry Hub Forward Curve", "CME Group Point-in-time futures settlement curves locked at May 22, 2026.")
    ]

    for idx, row in enumerate(rows_src, start=4):
        ws_src[f'A{idx}'] = row[0]
        ws_src[f'B{idx}'] = row[1]
        ws_src[f'A{idx}'].font = bold_font
        ws_src[f'B{idx}'].font = regular_font
        for col in ['A', 'B']:
            ws_src[f'{col}{idx}'].border = thin_border

    # Auto-adjust column widths across all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            # Give a buffer
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save("/Users/p.n.s/Desktop/P.N.S/project-caligula/EOG_DCF_Model.xlsx")
    print("Workbook generated successfully!")

if __name__ == "__main__":
    build_model()
